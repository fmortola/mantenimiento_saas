from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


def crear_excel(titulo, columnas, datos, nombre_hoja="Datos"):
    """
    Crea un archivo Excel con formato profesional.

    Args:
        titulo: Título del reporte
        columnas: Lista de nombres de columnas
        datos: Lista de listas con los datos
        nombre_hoja: Nombre de la hoja

    Returns:
        BytesIO con el archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título del reporte
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Fecha de generación
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columnas))
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal="center")

    # Encabezados (fila 4)
    for col, columna in enumerate(columnas, 1):
        cell = ws.cell(row=4, column=col, value=columna)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Datos
    for row_idx, fila in enumerate(datos, 5):
        for col_idx, valor in enumerate(fila, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Ajustar ancho de columnas
    for col in range(1, len(columnas) + 1):
        max_length = len(str(columnas[col-1]))
        for row in range(5, len(datos) + 5):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Congelar encabezados
    ws.freeze_panes = 'A5'

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def exportar_ordenes(ordenes, titulo="Órdenes de Trabajo"):
    """Exportar órdenes de trabajo a Excel"""
    columnas = [
        "Número", "Fecha Creación", "Cliente", "Ubicación", "Tipo",
        "Descripción", "Estado", "Prioridad", "Técnicos",
        "Fecha Programada", "Fecha Inicio", "Fecha Fin"
    ]

    datos = []
    for orden in ordenes:
        tecnicos = ", ".join([t.nombre for t in orden.tecnicos])
        datos.append([
            orden.numero,
            orden.fecha_creacion.strftime('%d/%m/%Y %H:%M') if orden.fecha_creacion else "",
            orden.cliente.nombre if orden.cliente else orden.cliente_rapido_nombre or "",
            orden.ubicacion.nombre if orden.ubicacion else orden.cliente_rapido_direccion or "",
            orden.tipo,
            orden.descripcion_solicitud or "",
            orden.estado,
            orden.prioridad,
            tecnicos,
            orden.fecha_programada.strftime('%d/%m/%Y %H:%M') if orden.fecha_programada else "",
            orden.fecha_inicio.strftime('%d/%m/%Y %H:%M') if orden.fecha_inicio else "",
            orden.fecha_fin.strftime('%d/%m/%Y %H:%M') if orden.fecha_fin else ""
        ])

    return crear_excel(titulo, columnas, datos, "Ordenes")


def exportar_mantenimientos(mantenimientos, titulo="Mantenimientos"):
    """Exportar mantenimientos a Excel"""
    columnas = [
        "Número", "Fecha Creación", "Cliente", "Ubicación", "Tipo",
        "Título", "Estado", "Técnicos", "Equipos Total", "Equipos Completados",
        "Progreso %", "Fecha Programada", "Fecha Inicio", "Fecha Fin"
    ]

    datos = []
    for mant in mantenimientos:
        tecnicos = ", ".join([t.nombre for t in mant.tecnicos])
        total_equipos = mant.equipos_mantenimiento.count()
        completados = mant.equipos_con_mantenimiento()
        datos.append([
            mant.numero,
            mant.fecha_creacion.strftime('%d/%m/%Y %H:%M') if mant.fecha_creacion else "",
            mant.cliente.nombre,
            mant.ubicacion.nombre,
            mant.tipo,
            mant.titulo,
            mant.estado,
            tecnicos,
            total_equipos,
            completados,
            f"{mant.progreso_porcentaje()}%",
            mant.fecha_programada.strftime('%d/%m/%Y %H:%M') if mant.fecha_programada else "",
            mant.fecha_inicio.strftime('%d/%m/%Y %H:%M') if mant.fecha_inicio else "",
            mant.fecha_fin.strftime('%d/%m/%Y %H:%M') if mant.fecha_fin else ""
        ])

    return crear_excel(titulo, columnas, datos, "Mantenimientos")


def exportar_tickets(tickets, titulo="Tickets"):
    """Exportar tickets a Excel"""
    columnas = [
        "Número", "Fecha Creación", "Cliente", "Ubicación", "Asunto",
        "Descripción", "Estado", "Prioridad", "Técnicos",
        "Fecha Asignación", "Fecha Resolución"
    ]

    datos = []
    for ticket in tickets:
        tecnicos = ", ".join([t.nombre for t in ticket.tecnicos])
        datos.append([
            ticket.numero,
            ticket.fecha_creacion.strftime('%d/%m/%Y %H:%M') if ticket.fecha_creacion else "",
            ticket.cliente.nombre,
            ticket.ubicacion.nombre if ticket.ubicacion else "",
            ticket.asunto,
            ticket.descripcion or "",
            ticket.estado,
            ticket.prioridad,
            tecnicos,
            ticket.fecha_asignacion.strftime('%d/%m/%Y %H:%M') if ticket.fecha_asignacion else "",
            ticket.fecha_resolucion.strftime('%d/%m/%Y %H:%M') if ticket.fecha_resolucion else ""
        ])

    return crear_excel(titulo, columnas, datos, "Tickets")


def exportar_productividad_tecnicos(tecnicos_data, fecha_desde, fecha_hasta, titulo="Productividad por Técnico"):
    """Exportar reporte de productividad por técnico"""
    columnas = [
        "Técnico", "Órdenes Completadas", "Mantenimientos Completados",
        "Tickets Resueltos", "Total Trabajos", "Equipos Atendidos"
    ]

    datos = []
    for data in tecnicos_data:
        datos.append([
            data['nombre'],
            data['ordenes_completadas'],
            data['mantenimientos_completados'],
            data['tickets_resueltos'],
            data['total'],
            data['equipos_atendidos']
        ])

    periodo = f" ({fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')})"
    return crear_excel(titulo + periodo, columnas, datos, "Productividad")


def exportar_historial_cliente(cliente, ordenes, mantenimientos, tickets, titulo=None):
    """Exportar historial completo de un cliente"""
    if not titulo:
        titulo = f"Historial - {cliente.nombre}"

    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def crear_hoja(ws, nombre, columnas, datos):
        ws.title = nombre

        # Encabezados
        for col, columna in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col, value=columna)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # Datos
        for row_idx, fila in enumerate(datos, 2):
            for col_idx, valor in enumerate(fila, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = thin_border

        # Ajustar anchos
        for col in range(1, len(columnas) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        ws.freeze_panes = 'A2'

    # Hoja de Órdenes
    ws_ordenes = wb.active
    ordenes_data = []
    for o in ordenes:
        ordenes_data.append([
            o.numero, o.fecha_creacion.strftime('%d/%m/%Y') if o.fecha_creacion else "",
            o.tipo, o.estado, o.descripcion_solicitud or ""
        ])
    crear_hoja(ws_ordenes, "Órdenes", ["Número", "Fecha", "Tipo", "Estado", "Descripción"], ordenes_data)

    # Hoja de Mantenimientos
    ws_mant = wb.create_sheet("Mantenimientos")
    mant_data = []
    for m in mantenimientos:
        mant_data.append([
            m.numero, m.fecha_creacion.strftime('%d/%m/%Y') if m.fecha_creacion else "",
            m.tipo, m.estado, m.titulo, f"{m.progreso_porcentaje()}%"
        ])
    crear_hoja(ws_mant, "Mantenimientos", ["Número", "Fecha", "Tipo", "Estado", "Título", "Progreso"], mant_data)

    # Hoja de Tickets
    ws_tickets = wb.create_sheet("Tickets")
    tickets_data = []
    for t in tickets:
        tickets_data.append([
            t.numero, t.fecha_creacion.strftime('%d/%m/%Y') if t.fecha_creacion else "",
            t.asunto, t.estado, t.prioridad
        ])
    crear_hoja(ws_tickets, "Tickets", ["Número", "Fecha", "Asunto", "Estado", "Prioridad"], tickets_data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def exportar_equipos(equipos, titulo="Inventario de Equipos"):
    """Exportar inventario de equipos a Excel"""
    columnas = [
        "Cliente", "Ubicación", "Departamento", "Tipo", "Nombre",
        "Marca", "Modelo", "Serial", "Condición", "Fecha Registro"
    ]

    datos = []
    for equipo in equipos:
        datos.append([
            equipo.ubicacion.cliente.nombre,
            equipo.ubicacion.nombre,
            equipo.departamento or "",
            equipo.tipo,
            equipo.nombre or "",
            equipo.marca or "",
            equipo.modelo or "",
            equipo.serial or "",
            equipo.condicion or "",
            equipo.fecha_creacion.strftime('%d/%m/%Y') if equipo.fecha_creacion else ""
        ])

    return crear_excel(titulo, columnas, datos, "Equipos")
